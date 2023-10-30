import cv2
import face_recognition
import openpyxl
import os
from datetime import datetime

# Load known student data (similar to the previous example)
student_data = {
    "Waseem Sajjad": ["Students\\Waseem Sajjad\\image1.jpg", "Students\\Waseem Sajjad\\image2.jpg"],
    "Babu Inayat": ["Students\\Babu Inayat\\image1.jpg", "Students\\Babu Inayat\\image2.jpg"],
    "Sarfaraz Ahmed": ["Students\\Sarfaraz Ahmed\\image1.jpg", "Students\\Sarfaraz Ahmed\\image2.jpg"],
    "Zahid Hussain": ["Students\\Zahid Hussain\\image1.jpg", "Students\\Zahid Hussain\\image2.jpg"]
    # Add more students and their image paths as needed
}

# Load known face encodings
known_face_encodings = {}
for student_name, image_paths in student_data.items():
    student_face_encodings = []
    for image_path in image_paths:
        image = face_recognition.load_image_file(image_path)
        face_encoding = face_recognition.face_encodings(image)[0]
        student_face_encodings.append(face_encoding)
    known_face_encodings[student_name] = student_face_encodings

# Specify the absolute path for the Excel sheet
excel_file_path = r"D:\PrecticleTestEnv\Attendence Sys\attendance.xlsx"

# Create or load the Excel sheet
if not os.path.exists(excel_file_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.cell(row=1, column=1, value="Student Name")
    sheet.cell(row=1, column=2, value="Attendance Status")
    sheet.cell(row=1, column=3, value="Date")
    sheet.cell(row=1, column=4, value="Time")
    row_num = 2
else:
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active
    row_num = sheet.max_row + 1

# Get today's date
today_date = datetime.now().strftime("%Y-%m-%d")

# Initialize webcam
video_capture = cv2.VideoCapture(0)  # Use 0 for default camera, or specify another camera index

while True:
    ret, frame = video_capture.read()

    # Find face locations in the current frame
    face_locations = face_recognition.face_locations(frame)
    face_encodings = face_recognition.face_encodings(frame, face_locations)

    for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
        # Compare the detected face encoding to known student face encodings
        for student_name, student_face_encodings in known_face_encodings.items():
            face_distance = face_recognition.face_distance(student_face_encodings, face_encoding)
            # Check if the smallest face_distance (highest confidence) is less than 0.6 (corresponds to >90% confidence)
            if min(face_distance) < 0.6:
                # Check if the student is already marked present for today
                already_marked = False
                for row in sheet.iter_rows(min_row=2, max_row=row_num - 1, min_col=1, max_col=1):
                    for cell in row:
                        if cell.value == student_name:
                            date_cell = sheet.cell(row=cell.row, column=3).value
                            if date_cell == today_date:
                                already_marked = True
                                break
                if not already_marked:
                    # Mark the student as present in your attendance data with date and time
                    attendance_status = "Present"
                    date_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    sheet.cell(row=row_num, column=1, value=student_name)
                    sheet.cell(row=row_num, column=2, value=attendance_status)
                    sheet.cell(row=row_num, column=3, value=today_date)
                    sheet.cell(row=row_num, column=4, value=date_time.split()[1])
                    row_num += 1
                    print(f"{student_name} is present!")

    # Display the video feed (you can comment this out if you don't want to see the video)
    cv2.imshow('Video', frame)
    
    # Close the OpenCV window when the 'Escape' key is pressed (ASCII code 27)
    if cv2.waitKey(1) == 27:
        break

# Save the attendance data to the Excel file
workbook.save(excel_file_path)

# Release the webcam and close the OpenCV window
video_capture.release()
cv2.destroyAllWindows()